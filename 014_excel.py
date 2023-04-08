import openpyxl

# wb =  open.Workbook()?
wb = openpyxl.load_workbook('transactions.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']

cell = sheet['a1']
print(cell.row)
print(cell.column)
print(cell.coordinate)

sheet.cell(row=1,column=1)

